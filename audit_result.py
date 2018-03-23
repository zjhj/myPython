#! /usr/bin/env python
#  -*- coding: utf-8 -*-
# collect information from sheets contain idc data

import sys
import traceback
import datetime
import re

from openpyxl.reader.excel import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.reader.excel import InvalidFileException

from nmap import PortScanner

try:
	if len( sys.argv ) != 2:
		raise Exception( 'Usage: %s miit_export_file.' % sys.argv[0] )
	# print sys.argv,len( sys.argv )

	wb = load_workbook( sys.argv[1] )
	sheets = wb.get_sheet_names()
	wb.close()

	p = re.compile( '^(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|[1-9])\\.(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|\\d)\\.(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|\\d)\\.(1\\d{2}|2[0-4]\\d|25[0-5]|[1-9]\\d|\\d)$' )

	# set default paramters
	ipcol_s = 2
	ipcol_e = 3

	port_min = 1
	port_max = 65535

	ip1_num_min = 0
	ip1_num_max = 255

	# np = PortScanner()
	port_scope = str(port_min) + '-' + str(port_max)

	n_rec = 0
	for curr_sheet in sheets:
		ws = wb.get_sheet_by_name( curr_sheet )

		for rn in range( ws.max_row ):
			if n_rec >= 10:
				break

			curr_ip_s = ws.cell( row=rn+1,column=ipcol_s ).value
			curr_ip_e = ws.cell( row=rn+1,column=ipcol_e ).value
			# check the validity of ip datas
			if p.match( curr_ip_s ) == None or p.match( curr_ip_e ) == None:
				continue

			n_rec += 1

			if curr_ip_s == curr_ip_e:
				np = PortScanner()
				scan_res = np.scan( curr_ip_s,port_scope )
				print scan_res
				# print curr_ip_s
			else:
				ip1_list = curr_ip_s.split( '.' )
				ip2_list = curr_ip_e.split( '.' )

				ip1_num_1st = int( ip1_list[0] )
				ip1_num_2nd = int( ip1_list[1] )
				ip1_num_3rd = int( ip1_list[2] )
				ip1_num_4th = int( ip1_list[3] )

				ip2_num_1st = int( ip2_list[0] )
				ip2_num_2nd = int( ip2_list[1] )
				ip2_num_3rd = int( ip2_list[2] )
				ip2_num_4th = int( ip2_list[3] )

				# increase the ip-number, until bigger than curr_ip_e
				while True:
					curr_ip = str(ip1_num_1st) + '.' + str(ip1_num_2nd) + '.' + str(ip1_num_3rd) + '.' + str(ip1_num_4th)
					if not ( ip1_num_1st <= ip2_num_1st and ip1_num_2nd <= ip2_num_2nd and ip1_num_3rd <= ip2_num_3rd and ip1_num_4th <= ip2_num_4th ):
						break

					np = PortScanner()
					scan_res = np.scan( curr_ip,port_scope )
					print scan_res
					# print curr_ip

					if ip1_num_4th >= ip1_num_max:
						if ip1_num_3rd >= ip1_num_max:
							if ip1_num_2nd >= ip1_num_max:
								if ip1_num_1st >= ip1_num_max:
									raise Exception( 'IP overflow' )
								else:
									ip1_num_1st += 1
									ip1_num_2nd = ip1_num_min
									ip1_num_3rd = ip1_num_min
									ip1_num_4th = ip1_num_min
							else:
								ip1_num_2nd += 1
								ip1_num_3rd = ip1_num_min
								ip1_num_4th = ip1_num_min
						else:
							ip1_num_4th = ip1_num_min
							ip1_num_3rd += 1
					else:
						ip1_num_4th += 1




		"""
		for rn in range(ws.max_row):
			curr_ip = ws.cell( row=rn+1,column=1 ).value
			curr_name = ws.cell( row=rn+1,column=2 ).value
			curr_duty = ws.cell( row=rn+1,column=3 ).value
			if curr_ip != None and p.match( curr_ip ) != None:
				if curr_name != None:
					if curr_duty != None:
						# print curr_ip + "," + curr_name + "," + str(curr_duty)
						print curr_ip , curr_name , curr_duty
						# print type(curr_ip) , type(curr_name) , type(curr_duty)
					else:
						print curr_ip , curr_name
		"""

	print n_rec

except IOError:
	print "Fail to open specified file!"
except InvalidFileException:
	print "Excel 2007 and aboves are expected!"
except Exception as e:
	print e
	traceback.print_exc()
